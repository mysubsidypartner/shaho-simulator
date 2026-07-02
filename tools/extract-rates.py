#!/usr/bin/env python3
"""Excel 料率表 → rates-data.json / rates.js 変換スクリプト"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)

FILES = {
    "tokyo_over40": ("東京都", True),
    "tokyo_under40": ("東京都", False),
    "osaka_over40": ("大阪府", True),
    "osaka_under40": ("大阪府", False),
}


def parse_max_pay(d_max):
    if isinstance(d_max, (int, float)):
        return int(d_max)
    if d_max and "未満" in str(d_max):
        m = re.search(r"([\d,]+)", str(d_max).split("\n")[-1])
        if m:
            return int(m.group(1).replace(",", "")) - 1
    return 10**9


def _num_or_none(value):
    if isinstance(value, (int, float)):
        return value
    return None


def parse_grades(path, with_care):
    ws = openpyxl.load_workbook(path, data_only=True).active
    grades = []
    for r in range(7, 51):
        grade = ws.cell(r, 2).value
        if grade is None:
            continue
        health_col = 9 if with_care else 7
        health_half_col = 10 if with_care else 8
        grades.append(
            {
                "grade": str(grade),
                "standardMonthly": int(ws.cell(r, 3).value),
                "minPay": int(ws.cell(r, 4).value),
                "maxPay": parse_max_pay(ws.cell(r, 6).value),
                "healthFull": ws.cell(r, health_col).value,
                "healthHalf": ws.cell(r, health_half_col).value,
                "childFull": ws.cell(r, 11).value,
                "childHalf": ws.cell(r, 12).value,
                "pensionFull": _num_or_none(ws.cell(r, 13).value),
                "pensionHalf": _num_or_none(ws.cell(r, 14).value),
            }
        )
    rate_col = "I5" if with_care else "G5"
    rates = {
        "health": float(str(ws[rate_col].value)),
        "child": float(ws["K5"].value),
        "pension": 0.183,
    }
    return {"grades": grades, "rates": rates}


def main():
    if len(sys.argv) != 5:
        print(
            "Usage: extract-rates.py "
            "tokyo_over40.xlsx tokyo_under40.xlsx osaka_over40.xlsx osaka_under40.xlsx"
        )
        sys.exit(1)

    paths = list(zip(FILES.keys(), sys.argv[1:5]))
    all_data = {}
    for key, path in paths:
        _, with_care = FILES[key]
        all_data[key] = parse_grades(path, with_care)
        print(f"{key}: {len(all_data[key]['grades'])} grades, rates={all_data[key]['rates']}")

    root = Path(__file__).resolve().parent.parent
    json_path = root / "public" / "js" / "rates-data.json"
    js_path = root / "public" / "js" / "rates.js"

    json_path.write_text(json.dumps(all_data, ensure_ascii=False, indent=2), encoding="utf-8")

    js = "// Auto-generated from Excel rate tables (令和8年度)\n"
    js += "const RATE_TABLES = " + json.dumps(all_data, ensure_ascii=False, indent=2) + ";\n\n"
    js += """
function getRateTable(location, ageCategory) {
  const key = location + '_' + (ageCategory === '40歳以上' ? 'over40' : 'under40');
  const map = {
    '東京都_over40': 'tokyo_over40',
    '東京都_under40': 'tokyo_under40',
    '大阪府_over40': 'osaka_over40',
    '大阪府_under40': 'osaka_under40',
  };
  const tableKey = map[key];
  if (!tableKey || !RATE_TABLES[tableKey]) {
    throw new Error('料率テーブルが見つかりません: ' + key);
  }
  return RATE_TABLES[tableKey];
}
"""
    js_path.write_text(js, encoding="utf-8")
    print(f"Wrote {json_path}")
    print(f"Wrote {js_path}")


if __name__ == "__main__":
    main()
